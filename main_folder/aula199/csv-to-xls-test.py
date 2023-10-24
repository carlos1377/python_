# type: ignore
import openpyxl
import csv
from pathlib import Path

DESKTOP_FOLDER = Path.home() / 'Desktop'
CSV_PATH = DESKTOP_FOLDER / 'teste.csv'
XLSX_PATH = DESKTOP_FOLDER / 'teste.xlsx'

# workbook_xl = openpyxl.load_workbook(XLSX_PATH)


with open(CSV_PATH, 'r', encoding='utf8') as cfile:
    reader = csv.reader(cfile, delimiter=';')
    wb = openpyxl.Workbook()
    ws = wb.active
    for chave, valor in reader:
        ws.append((chave, valor))

    for col_cell in ws['B']:
        col_cell.number_format = "R$ 0.00"

    ws['B1'].number_format = "@"

    wb.save(XLSX_PATH)
